
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from Evaluacion.views.info_db import obtener_db_info
from ..models import EvaluacionEstudiante, EvaluacionDirectivo, EvaluacionDocente
from django.shortcuts import render
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook

@login_required
def vista_reportes_evaluaciones(request):
    contexto = obtener_db_info(request)
    return render(request, 'core/reportes.html', contexto)

@login_required
def exportar_evaluaciones_estudiantes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluaciones Estudiantes"

    columnas = ['Estudiante', 'Materia', 'Periodo', 'Fecha Respuesta', 'Respuestas']
    ws.append(columnas)

    for evaluacion in EvaluacionEstudiante.objects.select_related('estudiante', 'materia', 'periodo'):
        respuestas_texto = ", ".join([f"{k}: {v}" for k, v in evaluacion.respuestas.items()])
        fila = [
            str(evaluacion.estudiante),
            str(evaluacion.materia),
            str(evaluacion.periodo),
            evaluacion.fecha_respuesta.strftime("%Y-%m-%d %H:%M"),
            respuestas_texto,
        ]
        ws.append(fila)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=evaluaciones_estudiantes.xlsx'
    return response

@login_required
def exportar_evaluaciones_docentes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluaciones Docentes"

    columnas = ['Docente', 'Periodo', 'Fecha Respuesta', 'Respuestas']
    ws.append(columnas)

    for evaluacion in EvaluacionDocente.objects.select_related('docente', 'periodo'):
        respuestas_texto = ", ".join([f"{k}: {v}" for k, v in evaluacion.respuestas.items()])
        fila = [
            str(evaluacion.docente),
            str(evaluacion.periodo),
            evaluacion.fecha_respuesta.strftime("%Y-%m-%d %H:%M"),
            respuestas_texto,
        ]
        ws.append(fila)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=evaluaciones_docentes.xlsx'
    return response

@login_required
def exportar_evaluaciones_directivos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluaciones Directivos"

    columnas = ['Evaluador', 'Docente Evaluado', 'Periodo', 'Fecha Respuesta', 'Respuestas']
    ws.append(columnas)

    for evaluacion in EvaluacionDirectivo.objects.select_related('evaluador', 'docente_evaluado', 'periodo'):
        respuestas_texto = ", ".join([f"{k}: {v}" for k, v in evaluacion.respuestas.items()])
        fila = [
            str(evaluacion.evaluador),
            str(evaluacion.docente_evaluado),
            str(evaluacion.periodo),
            evaluacion.fecha_respuesta.strftime("%Y-%m-%d %H:%M"),
            respuestas_texto,
        ]
        ws.append(fila)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=evaluaciones_directivos.xlsx'
    return response