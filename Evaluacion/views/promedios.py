from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from home.models.carga_academica.datos_adicionales import Programa, ProgramaUser
from Evaluacion.models import EvaluacionEstudiante, EvaluacionDocente, EvaluacionDirectivo
from home.models.talento_humano.usuarios import Empleado, EmpleadoUser
from home.models.carga_academica.carga_academica import CargaAcademica
from Evaluacion.views.info_db import obtener_db_info
import openpyxl


PONDERACIONES = {
    "estudiantes": 0.4,
    "directivos": 0.4,
    "autoevaluacion": 0.2,
}

def obtener_calificaciones_estudiantes_por_docente(programa_id=None):
    evaluaciones = EvaluacionEstudiante.objects.select_related('materia')
    resultados = []
    for evaluacion in evaluaciones:
        if programa_id and evaluacion.materia.fk_programa_id != int(programa_id):
            continue
        respuestas = evaluacion.respuestas
        if respuestas:
            calificaciones = [int(v) for v in respuestas.values()]
            promedio = sum(calificaciones) / len(calificaciones) if calificaciones else 0
            carga_academica = evaluacion.materia.cargaacademica_set.first()
            if carga_academica:
                resultados.append({
                    'docente': carga_academica.fk_docente_asignado.id,
                    'promedio_calificacion': promedio,
                })
    return resultados

def obtener_calificaciones_directivo(programa_id):
    cargas = CargaAcademica.objects.filter(fk_programa_id=programa_id)
    docentes_ids = cargas.values_list('fk_docente_asignado_id', flat=True).distinct()
    directivos_qs = EvaluacionDirectivo.objects.filter(docente_evaluado_id__in=docentes_ids)
    resultados = []
    for evaluacion in directivos_qs:
        respuestas = evaluacion.respuestas
        if respuestas:
            calificaciones = [int(v) for v in respuestas.values()]
            promedio = sum(calificaciones) / len(calificaciones) if calificaciones else 0
        else:
            promedio = 0
        resultados.append({
            'docente_evaluado': evaluacion.docente_evaluado.id,
            'promedio_calificacion': promedio,
        })
    return resultados

def obtener_calificaciones_autoevaluacion_docente(programa_id=None):
    evaluaciones = EvaluacionDocente.objects.select_related('docente')
    resultados = []

    for evaluacion in evaluaciones:
       
        empleado_user = EmpleadoUser.objects.filter(fk_user=evaluacion.docente).select_related('fk_empleado').first()

        if not empleado_user:
            continue

       
        programa_asignado = ProgramaUser.objects.filter(
            fk_user=evaluacion.docente, 
            fk_programa__id=programa_id
        ).exists()

        if programa_id and not programa_asignado:
            continue

    
        respuestas = evaluacion.respuestas
        valores = [int(v) for v in respuestas.values() if isinstance(v, (int, float))]
        promedio_calificacion = sum(valores) / len(valores) if valores else 0

        
        resultados.append({
            'docente_id': empleado_user.fk_empleado.id,  
            'docente_nombre': f"{empleado_user.fk_empleado.primer_nombre} {empleado_user.fk_empleado.primer_apellido}",
            'promedio_calificacion': promedio_calificacion,
        })

    return resultados

def categorizar_desempeno(promedio):
    if promedio >= 4.5:
        return "Excelente"
    elif promedio >= 4.0:
        return "Bueno"
    elif promedio >= 3.5:
        return "Aceptable"
    else:
        return "Bajo"

def calcular_desempeno_docentes_categorizado(programa_id=None):
    estudiantes = obtener_calificaciones_estudiantes_por_docente(programa_id)
    directivos = obtener_calificaciones_directivo(programa_id)
    autoevaluacion = obtener_calificaciones_autoevaluacion_docente(programa_id)
    resultados = {}
    for item in estudiantes:
        docente_id = item['docente']
        promedio = item['promedio_calificacion']
        if docente_id not in resultados:
            resultados[docente_id] = {'estudiantes': 0, 'directivos': 0, 'autoevaluacion': 0}
        resultados[docente_id]['estudiantes'] = promedio
    for item in directivos:
        docente_id = item['docente_evaluado']
        promedio = item['promedio_calificacion']
        if docente_id not in resultados:
            resultados[docente_id] = {'estudiantes': 0, 'directivos': 0, 'autoevaluacion': 0}
        resultados[docente_id]['directivos'] = promedio
    for item in autoevaluacion:
        docente_id = item['docente_id']
        promedio = item['promedio_calificacion']
        if docente_id not in resultados:
            resultados[docente_id] = {'estudiantes': 0, 'directivos': 0, 'autoevaluacion': 0}
        resultados[docente_id]['autoevaluacion'] = promedio
    desempeno = []
    for docente_id, promedios in resultados.items():
        promedio_ponderado = (
            promedios['estudiantes'] * PONDERACIONES['estudiantes'] +
            promedios['directivos'] * PONDERACIONES['directivos'] +
            promedios['autoevaluacion'] * PONDERACIONES['autoevaluacion']
        )
        desempeno.append({
            'docente_id': docente_id,
            'promedio': promedio_ponderado,
            'desempeño': categorizar_desempeno(promedio_ponderado),
        })
    return desempeno

def agregar_nombre_docente(lista, key_id):
    for item in lista:
        if isinstance(item, dict):
            docente_id = item.get(key_id)
            usuario = Empleado.objects.filter(id=docente_id).first() if docente_id else None
            item['docente_nombre'] = str(usuario) if usuario else "Desconocido"
            item['docente_id'] = docente_id if docente_id else "N/A"
        else:
            docente_obj = getattr(item, key_id, None)
            docente_id = docente_obj.id if docente_obj else None
            usuario = Empleado.objects.filter(id=docente_id).first() if docente_id else None
            setattr(item, 'docente_nombre', str(usuario) if usuario else "Desconocido")
            setattr(item, 'docente_id', docente_id if docente_id else "N/A")
    return lista

@login_required
def desempeno_por_programa(request):
    programas = Programa.objects.all()
    programa_seleccionado = request.GET.get('programa', None)
    tipo_informe = request.GET.get('funcion', None)
    estudiantes = []
    directivos = []
    autoevaluacion = []
    desempeno = []
    if programa_seleccionado and tipo_informe:
        if tipo_informe == 'estudiantes':
            estudiantes = obtener_calificaciones_estudiantes_por_docente(programa_seleccionado)
            estudiantes = agregar_nombre_docente(estudiantes, 'docente')
        elif tipo_informe == 'directivos':
            directivos = obtener_calificaciones_directivo(programa_seleccionado)
            directivos = agregar_nombre_docente(directivos, 'docente_evaluado')
        elif tipo_informe == 'autoevaluacion':
            autoevaluacion = obtener_calificaciones_autoevaluacion_docente(programa_seleccionado)
            autoevaluacion = agregar_nombre_docente(autoevaluacion, 'docente_id')
        elif tipo_informe == 'desempeno':
            desempeno = calcular_desempeno_docentes_categorizado(programa_seleccionado)
            desempeno = agregar_nombre_docente(desempeno, 'docente_id')
    contexto = obtener_db_info(request)
    contexto.update({
        'programas': programas,
        'programa_seleccionado': programa_seleccionado,
        'funcion': tipo_informe,
        'estudiantes': estudiantes,
        'directivos': directivos,
        'autoevaluacion': autoevaluacion,
        'desempeno': desempeno,
    })
    return render(request, 'core/promedios_docentes.html', contexto)

@login_required
def exportar_informe_excel(request):
    programa_seleccionado = request.GET.get('programa', None)
    tipo_informe = request.GET.get('tipo_informe', None)

    if tipo_informe not in ["detallado", "general"]:
        return HttpResponse("Tipo de informe no válido", status=400)

    
    workbook = openpyxl.Workbook()

    if tipo_informe == "detallado":
        
        sheet1 = workbook.active
        sheet1.title = "Calificaciones Estudiantes"
        encabezados = ["ID", "Docente", "Promedio", "Desempeño"]
        data = obtener_calificaciones_estudiantes_por_docente(programa_seleccionado)
        data = agregar_nombre_docente(data, "docente")
        sheet1.append(encabezados)
        for item in data:
            sheet1.append([
                item.get("docente_id", "Desconocido"),
                item.get("docente_nombre", "Desconocido"),
                item.get("promedio_calificacion", 0),
                categorizar_desempeno(item.get("promedio_calificacion", 0)),
            ])

        sheet2 = workbook.create_sheet(title="Calificaciones Directivos")
        data = obtener_calificaciones_directivo(programa_seleccionado)
        data = agregar_nombre_docente(data, "docente_evaluado")
        sheet2.append(encabezados)
        for item in data:
            sheet2.append([
                item.get("docente_id", "Desconocido"),
                item.get("docente_nombre", "Desconocido"),
                item.get("promedio_calificacion", 0),
                categorizar_desempeno(item.get("promedio_calificacion", 0)),
            ])

        sheet3 = workbook.create_sheet(title="Autoevaluaciones Docentes")
        data = obtener_calificaciones_autoevaluacion_docente(programa_seleccionado)
        data = agregar_nombre_docente(data, "docente_id")
        sheet3.append(encabezados)
        for item in data:
            sheet3.append([
                item.get("docente_id", "Desconocido"),
                item.get("docente_nombre", "Desconocido"),
                item.get("promedio_calificacion", 0),
                categorizar_desempeno(item.get("promedio_calificacion", 0)),
            ])

        
        sheet4 = workbook.create_sheet(title="Desempeño General")
        data = calcular_desempeno_docentes_categorizado(programa_seleccionado)
        data = agregar_nombre_docente(data, "docente_id")
        sheet4.append(encabezados)
        for item in data:
            sheet4.append([
                item.get("docente_id", "Desconocido"),
                item.get("docente_nombre", "Desconocido"),
                item.get("promedio", 0),
                item.get("desempeño", "-"),
            ])

    elif tipo_informe == "general":
      
        sheet = workbook.active
        sheet.title = "Informe General"
        encabezados = ["ID", "Docente", "Promedio", "Desempeño"]
        data = calcular_desempeno_docentes_categorizado(programa_seleccionado)
        data = agregar_nombre_docente(data, "docente_id")
        sheet.append(encabezados)
        for item in data:
            sheet.append([
                item.get("docente_id", "Desconocido"),
                item.get("docente_nombre", "Desconocido"),
                item.get("promedio", 0),
                item.get("desempeño", "-"),
            ])

    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{tipo_informe}_informe.xlsx"'
    workbook.save(response)
    return response

